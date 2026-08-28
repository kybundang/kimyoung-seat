from fastapi import FastAPI, Form, UploadFile, File, Request
from fastapi.responses import HTMLResponse
import sqlite3
import pandas as pd
import io
import re
import json
import hashlib
from collections import Counter
from datetime import datetime
import os

app = FastAPI()

SECRET_KEY = "kimyoung_seat_secret_key"
DB_PATH = os.path.abspath("seat_system.db")

def get_today_str():
    return datetime.now().strftime("%Y%m%d")

def generate_daily_token(date_str: str):
    return hashlib.sha256(f"{date_str}_{SECRET_KEY}".encode()).hexdigest()[:12]

def get_db_connection():
    conn = sqlite3.connect(DB_PATH, timeout=15.0, isolation_level=None)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA busy_timeout=5000;")
    return conn

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS students (
            username TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            class_name TEXT DEFAULT '-',
            password TEXT DEFAULT '1234',
            status TEXT DEFAULT 'active'
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS rooms (
            room_name TEXT PRIMARY KEY,
            title TEXT,
            rows_count INTEGER,
            cols_count INTEGER,
            grid_json TEXT
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS class_configs (
            class_name TEXT PRIMARY KEY,
            room_name TEXT DEFAULT '301호',
            rows_count INTEGER DEFAULT 0,
            cols_count INTEGER DEFAULT 0,
            reset_datetime TEXT DEFAULT '',
            last_reset_at TEXT DEFAULT ''
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS seat_reservations (
            class_name TEXT,
            seat_id TEXT,
            username TEXT,
            name TEXT,
            PRIMARY KEY (class_name, seat_id)
        )
    """)
    
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_students_class ON students(class_name);")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_reservations_class ON seat_reservations(class_name);")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_reservations_user ON seat_reservations(username);")
    
    conn.close()

init_db()

def check_and_apply_resets(conn):
    cursor = conn.cursor()
    now = datetime.now()
    now_str = now.strftime("%Y-%m-%dT%H:%M")
    today_date_str = now.strftime("%Y-%m-%d")
    current_hm = now.strftime("%H:%M")
    current_weekday = str(now.weekday())
    current_day = str(now.day)
    current_year_month = now.strftime("%Y-%m")
    current_year_week = f"{now.year}-W{now.isocalendar()[1]}"
    
    cursor.execute("SELECT class_name, reset_datetime, last_reset_at FROM class_configs")
    configs = cursor.fetchall()
    
    for c_name, r_dt, l_reset in configs:
        if not r_dt:
            continue
            
        should_reset = False
        new_reset_marker = ""

        if r_dt.startswith("DAILY:"):
            target_hm = r_dt.replace("DAILY:", "").strip()
            if current_hm >= target_hm and (not l_reset or l_reset < today_date_str):
                should_reset = True
                new_reset_marker = today_date_str

        elif r_dt.startswith("WEEKLY:"):
            parts = r_dt.replace("WEEKLY:", "").strip().split(":")
            if len(parts) == 3:
                target_wday, target_h, target_m = parts
                target_hm = f"{target_h}:{target_m}"
                if current_weekday == target_wday and current_hm >= target_hm and (not l_reset or l_reset < current_year_week):
                    should_reset = True
                    new_reset_marker = current_year_week

        elif r_dt.startswith("MONTHLY:"):
            parts = r_dt.replace("MONTHLY:", "").strip().split(":")
            if len(parts) == 3:
                target_day, target_h, target_m = parts
                target_hm = f"{target_h}:{target_m}"
                if current_day == target_day and current_hm >= target_hm and (not l_reset or l_reset < current_year_month):
                    should_reset = True
                    new_reset_marker = current_year_month

        elif "T" in r_dt:
            if r_dt <= now_str and (not l_reset or l_reset < r_dt):
                should_reset = True
                new_reset_marker = r_dt

        if should_reset:
            cursor.execute("DELETE FROM seat_reservations WHERE class_name = ?", (c_name,))
            cursor.execute("UPDATE class_configs SET last_reset_at = ? WHERE class_name = ?", (new_reset_marker, c_name))

def parse_student_file(contents: bytes) -> pd.DataFrame:
    try:
        df = pd.read_excel(io.BytesIO(contents))
        if '아이디' in df.columns and '이름' in df.columns: return df
    except Exception: pass

    text = ""
    for enc in ['utf-8-sig', 'euc-kr', 'cp949', 'utf-8']:
        try:
            text = contents.decode(enc)
            break
        except Exception: continue
    if not text: text = contents.decode('euc-kr', errors='ignore')

    if "<table" in text.lower() or "<tr" in text.lower():
        try:
            rows = []
            tr_matches = re.findall(r'<tr[^>]*>(.*?)</tr>', text, re.DOTALL | re.IGNORECASE)
            for tr in tr_matches:
                cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', tr, re.DOTALL | re.IGNORECASE)
                clean = [re.sub(r'<[^>]+>', '', c).replace('&nbsp;', ' ').strip() for c in cells]
                if any(clean): rows.append(clean)
            if rows:
                header, data = rows[0], rows[1:]
                max_l = max(len(r) for r in rows)
                header += [f"Col_{i}" for i in range(len(header), max_l)]
                data = [r + [''] * (max_l - len(r)) for r in data]
                df = pd.DataFrame(data, columns=header)
                if '아이디' in df.columns and '이름' in df.columns: return df
        except Exception: pass

    for sep in [',', '\t', '|']:
        try:
            df = pd.read_csv(io.StringIO(text), sep=sep)
            if '아이디' in df.columns and '이름' in df.columns: return df
        except Exception: pass

    return None

# ==============================================================================
# 정밀 좌석표 엑셀 파서
# ==============================================================================
def parse_seat_excel(contents: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    seat_pattern = re.compile(r'^([A-Za-z])(\d{1,2})$')
    
    parsed_rooms = []
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        max_r, max_c = ws.max_row, ws.max_column
        if max_r < 1 or max_c < 1:
            continue

        title = sheetname.strip()
        for r in range(1, min(max_r + 1, 6)):
            for c in range(1, max_c + 1):
                val = str(ws.cell(r, c).value or '').strip()
                if val and ('[' in val or '호' in val or '강의실' in val or '관' in val):
                    title = val
                    break
            if title != sheetname.strip():
                break

        seat_rows_set = set()
        seat_cols_set = set()
        distinct_seat_alphas = set()
        distinct_seat_nums = set()
        pure_seat_count = 0

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                val = str(ws.cell(r, c).value or '').strip().upper()
                match = seat_pattern.match(val)
                if match:
                    seat_rows_set.add(r)
                    seat_cols_set.add(c)
                    distinct_seat_alphas.add(match.group(1))
                    distinct_seat_nums.add(int(match.group(2)))
                    pure_seat_count += 1

        if not seat_rows_set:
            continue

        sorted_seat_rows = sorted(list(seat_rows_set))
        min_col = min(seat_cols_set)
        max_col = max(seat_cols_set)

        grid = []
        for r in sorted_seat_rows:
            row_cells = []
            for c in range(min_col, max_col + 1):
                val = str(ws.cell(r, c).value or '').strip().upper()
                if seat_pattern.match(val):
                    row_cells.append({"type": "seat", "id": val})
                elif val.isdigit() and len(val) <= 2:
                    row_cells.append({"type": "aisle", "val": val})
                else:
                    row_cells.append({"type": "empty"})
            grid.append(row_cells)

        pure_rows_count = len(distinct_seat_nums) if distinct_seat_nums else len(grid)
        pure_cols_count = len(distinct_seat_alphas) if distinct_seat_alphas else (max_col - min_col + 1)

        parsed_rooms.append({
            "room_name": sheetname.strip(),
            "title": title,
            "rows_count": pure_rows_count,
            "cols_count": pure_cols_count,
            "total_seat_count": pure_seat_count,
            "grid": grid
        })
        
    return parsed_rooms

# ==============================================================================
# 1. 학생용 접속 화면
# ==============================================================================
@app.get("/", response_class=HTMLResponse)
def student_view():
    return """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>김영편입 좌석표 신청</title>
        <script src="https://cdn.tailwindcss.com"></script>
        <style>
            .x-box {
                background: linear-gradient(to top right, transparent calc(50% - 1px), #cbd5e1, transparent calc(50% + 1px)),
                            linear-gradient(to bottom right, transparent calc(50% - 1px), #cbd5e1, transparent calc(50% + 1px));
            }
        </style>
    </head>
    <body class="bg-slate-100 flex justify-center p-3 sm:p-4 min-h-screen items-center">
        <div id="main-card" class="w-full max-w-md bg-white rounded-2xl p-5 sm:p-6 shadow-md border border-slate-200 transition-all duration-300">
            
            <div class="mb-4 bg-slate-50 border border-slate-200 rounded-xl p-3 text-center">
                <div id="student-live-date" class="text-xs font-bold text-slate-500"></div>
                <div id="student-live-clock" class="text-xl font-extrabold text-blue-600 tracking-wider mt-0.5 font-mono">00:00:00</div>
            </div>

            <div id="step-login">
                <div class="text-center mb-6">
                    <span class="inline-block bg-blue-100 text-blue-700 text-xs px-2.5 py-1 rounded-full font-bold mb-2">실시간 좌석신청</span>
                    <h2 class="text-xl font-black text-slate-900">김영편입 좌석표 신청</h2>
                </div>
                <div class="space-y-4">
                    <div>
                        <label class="block text-xs font-bold text-slate-600 mb-1">아이디</label>
                        <input type="text" id="username" placeholder="아이디 입력" class="w-full border rounded-lg p-2.5 text-sm outline-none focus:border-blue-500 font-medium">
                    </div>
                    <div>
                        <label class="block text-xs font-bold text-slate-600 mb-1">비밀번호</label>
                        <input type="password" id="password" placeholder="초기 비밀번호: 1234" class="w-full border rounded-lg p-2.5 text-sm outline-none focus:border-blue-500 font-medium" onkeydown="if(event.key==='Enter') login()">
                    </div>
                    <button onclick="login()" class="w-full bg-blue-600 text-white font-bold py-3 rounded-lg hover:bg-blue-700 transition shadow-sm text-base">로그인 하기</button>
                </div>
            </div>

            <div id="step-class" class="hidden">
                <div class="flex justify-between items-center bg-blue-50 p-3 rounded-xl mb-4">
                    <div>
                        <span id="user-info-text" class="text-sm font-bold text-blue-900"></span>
                        <p class="text-xs text-blue-600 font-medium">신청할 반을 선택하세요.</p>
                    </div>
                    <div class="flex gap-1.5">
                        <button onclick="openPwModal()" class="text-xs bg-white border border-slate-300 text-slate-700 px-2 py-1.5 rounded-lg hover:bg-slate-50 font-bold">비번변경</button>
                        <button onclick="logout()" class="text-xs bg-red-500 text-white px-2 py-1.5 rounded-lg hover:bg-red-600 font-bold">로그아웃</button>
                    </div>
                </div>
                <div class="text-xs font-bold text-slate-500 mb-2">수강 중인 반 목록</div>
                <div id="class-list-container" class="space-y-2 my-3"></div>
            </div>

            <div id="step-seat" class="hidden">
                <div class="flex justify-between items-center bg-slate-50 border p-3 rounded-xl mb-4">
                    <div>
                        <div id="selected-class-title" class="text-sm font-bold text-slate-800"></div>
                        <div id="room-info-badge" class="text-xs text-blue-600 font-bold mt-0.5"></div>
                    </div>
                    <button onclick="backToClassSelect()" class="text-xs bg-slate-200 text-slate-700 px-3 py-1.5 rounded-lg hover:bg-slate-300 font-bold">◀ 반 목록</button>
                </div>

                <div class="grid grid-cols-1 lg:grid-cols-12 gap-5 items-start">
                    <div class="lg:col-span-5 bg-white border border-slate-300 rounded-xl overflow-hidden shadow-sm">
                        <div class="bg-amber-400 p-2.5 text-white font-extrabold text-sm sm:text-base border-b border-amber-500">
                            <span id="excel-blueprint-title">[강의실]</span>
                        </div>
                        <div class="flex bg-slate-200 text-xs font-bold text-slate-700 border-b border-slate-300">
                            <div class="flex-1 py-1.5 text-center tracking-widest">칠 판</div>
                            <div class="bg-slate-500 text-white px-3 py-1.5 font-bold">출입문</div>
                        </div>
                        <div id="excel-blueprint-grid" class="p-2.5 max-h-[50vh] overflow-auto bg-white flex justify-center"></div>
                        <div class="bg-slate-50 p-2 border-t border-slate-200 text-[11px] text-slate-600 flex justify-around font-bold">
                            <span><b class="text-emerald-600">■</b> 내 좌석</span>
                            <span><b class="text-red-500">■</b> 신청 완료</span>
                            <span><b class="text-slate-400">■</b> 통로/빈자리</span>
                        </div>
                    </div>

                    <div class="lg:col-span-7">
                        <div class="w-full bg-slate-700 text-white py-1.5 rounded-lg text-xs font-bold text-center mb-2 tracking-widest shadow-sm">칠 판 / 교 탁 (앞면)</div>
                        <div id="seats-grid-container" class="grid gap-1.5 max-h-[50vh] overflow-auto p-2 bg-slate-50 rounded-xl border border-slate-200"></div>
                    </div>
                </div>
            </div>

        </div>

        <div id="pw-modal" class="fixed inset-0 bg-slate-900/50 hidden items-center justify-center z-50 p-4">
            <div class="bg-white rounded-2xl p-6 w-full max-w-sm shadow-xl">
                <h3 class="text-base font-bold text-slate-900 mb-4">비밀번호 변경</h3>
                <div class="space-y-3">
                    <input type="password" id="cur-pw" placeholder="현재 비밀번호" class="w-full border rounded-lg p-2 text-sm font-medium">
                    <input type="password" id="new-pw" placeholder="새 비밀번호" class="w-full border rounded-lg p-2 text-sm font-medium">
                    <input type="password" id="new-pw-confirm" placeholder="새 비밀번호 확인" class="w-full border rounded-lg p-2 text-sm font-medium">
                </div>
                <div class="flex justify-end gap-2 mt-5">
                    <button onclick="closePwModal()" class="px-3 py-2 border rounded-lg text-xs text-slate-600 font-bold">취소</button>
                    <button onclick="changePassword()" class="px-4 py-2 bg-blue-600 text-white rounded-lg text-xs font-bold shadow-sm">변경 완료</button>
                </div>
            </div>
        </div>

        <script>
            let currentUsername = "";
            let currentStudentName = "";
            let currentSelectedClass = "";
            let pollingInterval = null;

            function updateLiveClock() {
                const now = new Date();
                const days = ['일', '월', '화', '수', '목', '금', '토'];
                const year = now.getFullYear();
                const month = String(now.getMonth() + 1).padStart(2, '0');
                const date = String(now.getDate()).padStart(2, '0');
                const dayName = days[now.getDay()];
                
                const hours = String(now.getHours()).padStart(2, '0');
                const minutes = String(now.getMinutes()).padStart(2, '0');
                const seconds = String(now.getSeconds()).padStart(2, '0');
                
                const dateEl = document.getElementById("student-live-date");
                const clockEl = document.getElementById("student-live-clock");
                if (dateEl) dateEl.innerText = `${year}년 ${month}월 ${date}일 (${dayName})`;
                if (clockEl) clockEl.innerText = `${hours}:${minutes}:${seconds}`;
            }
            setInterval(updateLiveClock, 1000);
            updateLiveClock();

            async function login() {
                const uName = document.getElementById("username").value.trim();
                const pw = document.getElementById("password").value.trim();
                if(!uName || !pw) { alert("아이디와 비밀번호를 입력해주세요."); return; }

                const res = await fetch("/api/login", {
                    method: "POST",
                    headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                    body: `username=${encodeURIComponent(uName)}&password=${encodeURIComponent(pw)}`
                });
                const data = await res.json();
                if (data.success) {
                    currentUsername = data.username;
                    currentStudentName = data.name;
                    document.getElementById("user-info-text").innerText = `${currentStudentName} 님`;

                    const listContainer = document.getElementById("class-list-container");
                    listContainer.innerHTML = "";
                    data.class_list.forEach(c => {
                        listContainer.innerHTML += `
                            <button onclick="selectClass('${c}')" class="w-full text-left p-3.5 border border-slate-200 rounded-xl hover:border-blue-500 hover:bg-blue-50/50 transition flex justify-between items-center group shadow-xs">
                                <span class="text-sm font-bold text-slate-800 group-hover:text-blue-700">${c}</span>
                                <span class="text-xs bg-blue-100 text-blue-700 px-2.5 py-1 rounded-full font-bold">좌석 선택 ➔</span>
                            </button>
                        `;
                    });

                    document.getElementById("step-login").classList.add("hidden");
                    document.getElementById("step-class").classList.remove("hidden");
                } else {
                    alert(data.message);
                }
            }

            function selectClass(className) {
                currentSelectedClass = className;
                document.getElementById("selected-class-title").innerText = className;
                document.getElementById("main-card").className = "w-full max-w-6xl bg-white rounded-2xl p-5 sm:p-6 shadow-md border border-slate-200 transition-all duration-300";
                document.getElementById("step-class").classList.add("hidden");
                document.getElementById("step-seat").classList.remove("hidden");
                loadSeats();

                if (pollingInterval) clearInterval(pollingInterval);
                pollingInterval = setInterval(loadSeats, 1500);
            }

            function backToClassSelect() {
                if (pollingInterval) clearInterval(pollingInterval);
                document.getElementById("main-card").className = "w-full max-w-lg bg-white rounded-2xl p-5 sm:p-6 shadow-md border border-slate-200 transition-all duration-300";
                document.getElementById("step-seat").classList.add("hidden");
                document.getElementById("step-class").classList.remove("hidden");
            }

            function logout() {
                if (pollingInterval) clearInterval(pollingInterval);
                currentUsername = "";
                document.getElementById("main-card").className = "w-full max-w-md bg-white rounded-2xl p-5 sm:p-6 shadow-md border border-slate-200 transition-all duration-300";
                document.getElementById("step-class").classList.add("hidden");
                document.getElementById("step-seat").classList.add("hidden");
                document.getElementById("step-login").classList.remove("hidden");
            }

            async function loadSeats() {
                if(!currentSelectedClass || !currentUsername) return;
                try {
                    const res = await fetch(`/api/seats?username=${encodeURIComponent(currentUsername)}&class_name=${encodeURIComponent(currentSelectedClass)}`);
                    const data = await res.json();

                    document.getElementById("room-info-badge").innerText = `강의실: ${data.room_name} (${data.total_seats}석)`;
                    document.getElementById("excel-blueprint-title").innerText = data.room_title || `[${data.room_name}]`;

                    const bpGrid = document.getElementById("excel-blueprint-grid");
                    bpGrid.innerHTML = "";
                    const table = document.createElement("table");
                    table.className = "border-collapse text-center text-xs";
                    
                    data.grid.forEach(row => {
                        const tr = document.createElement("tr");
                        row.forEach(cell => {
                            const td = document.createElement("td");
                            if (cell.type === 'seat') {
                                let bg = "bg-white text-slate-800 border-slate-400";
                                if (cell.status_class === 'mine') bg = "bg-emerald-100 text-emerald-800 border-emerald-500 font-bold";
                                if (cell.status_class === 'occupied') bg = "bg-red-100 text-red-700 border-red-400";
                                
                                td.className = `border p-1 ${bg}`;
                                td.style.minWidth = "36px";
                                td.innerHTML = `<div class="font-bold">${cell.id}</div><div class="text-[9px] text-slate-400 font-normal">| &nbsp; |</div>`;
                            } else if (cell.type === 'aisle') {
                                td.className = "border-x border-dashed bg-slate-100/70 text-slate-500 font-bold px-1.5";
                                td.style.minWidth = "24px";
                                td.innerText = cell.val;
                            } else {
                                td.className = "border border-slate-200 x-box p-1 text-transparent";
                                td.style.minWidth = "36px";
                                td.innerText = "X";
                            }
                            tr.appendChild(td);
                        });
                        table.appendChild(tr);
                    });
                    bpGrid.appendChild(table);

                    const container = document.getElementById("seats-grid-container");
                    container.style.gridTemplateColumns = `repeat(${data.cols}, minmax(0, 1fr))`;
                    container.innerHTML = "";

                    data.grid.forEach(row => {
                        row.forEach(cell => {
                            const div = document.createElement("div");
                            if (cell.type === 'seat') {
                                let color = "bg-blue-100 text-blue-700 border-blue-200 hover:bg-blue-200 cursor-pointer shadow-xs active:scale-95";
                                if (cell.status_class === 'mine') color = "bg-emerald-500 text-white border-emerald-600 font-bold shadow-sm ring-2 ring-emerald-300";
                                if (cell.status_class === 'occupied') color = "bg-red-100 text-red-400 border-red-200 cursor-not-allowed opacity-60";

                                div.className = `p-2 rounded-lg border text-center text-xs font-bold transition flex items-center justify-center ${color}`;
                                div.style.minHeight = "38px";
                                div.innerText = cell.id;
                                if (cell.status_class !== 'occupied') {
                                    div.onclick = () => reserveSeat(cell.id);
                                }
                            } else if (cell.type === 'aisle') {
                                div.className = "flex items-center justify-center text-[11px] font-bold text-slate-400 bg-slate-200/50 rounded";
                                div.style.minHeight = "38px";
                                div.innerText = cell.val;
                            } else {
                                div.className = "p-2 opacity-0";
                                div.style.minHeight = "38px";
                            }
                            container.appendChild(div);
                        });
                    });
                } catch(e) {}
            }

            async function reserveSeat(seatId) {
                if (!confirm(`[${currentSelectedClass}] ${seatId}번 좌석으로 신청하시겠습니까?`)) return;
                const res = await fetch("/api/reserve", {
                    method: "POST",
                    headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                    body: `username=${encodeURIComponent(currentUsername)}&name=${encodeURIComponent(currentStudentName)}&class_name=${encodeURIComponent(currentSelectedClass)}&seat_id=${encodeURIComponent(seatId)}`
                });
                const data = await res.json();
                if (data.success) {
                    loadSeats();
                } else {
                    alert(data.message);
                    loadSeats();
                }
            }

            function openPwModal() { document.getElementById("pw-modal").classList.replace("hidden", "flex"); }
            function closePwModal() { document.getElementById("pw-modal").classList.replace("flex", "hidden"); }

            async function changePassword() {
                const curPw = document.getElementById("cur-pw").value;
                const newPw = document.getElementById("new-pw").value;
                const confirmPw = document.getElementById("new-pw-confirm").value;

                if (!curPw || !newPw) { alert("비밀번호를 입력하세요."); return; }
                if (newPw !== confirmPw) { alert("새 비밀번호 확인이 일치하지 않습니다."); return; }

                const res = await fetch("/api/change-password", {
                    method: "POST",
                    headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                    body: `username=${encodeURIComponent(currentUsername)}&current_password=${encodeURIComponent(curPw)}&new_password=${encodeURIComponent(newPw)}`
                });
                const data = await res.json();
                alert(data.message);
                if (data.success) closePwModal();
            }
        </script>
    </body>
    </html>
    """

# ==============================================================================
# 2. 관리자 화면
# ==============================================================================
@app.get("/admin/students", response_class=HTMLResponse)
def admin_students_view(request: Request):
    return """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>김영편입 좌석표 신청 - 관리자</title>
        <script src="https://cdn.tailwindcss.com"></script>
        <script src="https://cdn.jsdelivr.net/npm/qrcodejs@1.0.0/qrcode.min.js"></script>
    </head>
    <body class="bg-slate-50 flex flex-col md:flex-row h-screen text-slate-800 antialiased">
        
        <header id="mobile-header" class="md:hidden bg-white border-b border-slate-200 px-4 py-3 flex justify-between items-center sticky top-0 z-30 shadow-xs">
            <div class="flex items-center gap-3">
                <button onclick="toggleMobileDrawer()" class="p-2 rounded-lg bg-slate-100 hover:bg-slate-200 text-slate-800 font-bold transition">
                    <svg class="w-6 h-6" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 6h16M4 12h16M4 18h16"></path></svg>
                </button>
                <h1 class="text-base font-extrabold text-blue-700">김영편입 좌석관리</h1>
            </div>
            <div id="header-clock" class="text-xs font-mono font-bold text-slate-500"></div>
        </header>

        <div id="mobile-drawer" class="fixed inset-0 bg-slate-900/50 z-40 hidden md:hidden" onclick="toggleMobileDrawer()">
            <div class="w-64 bg-white h-full p-5 shadow-2xl flex flex-col justify-between" onclick="event.stopPropagation()">
                <div>
                    <div class="flex justify-between items-center mb-6 border-b pb-3">
                        <h2 class="text-lg font-black text-blue-700">메뉴 목록</h2>
                        <button onclick="toggleMobileDrawer()" class="text-slate-400 font-bold text-xl">&times;</button>
                    </div>
                    <nav class="space-y-2">
                        <button onclick="switchTab('students'); toggleMobileDrawer();" id="m-tab-students" class="w-full text-left px-3.5 py-2.5 bg-blue-50 text-blue-700 font-bold rounded-xl transition">학생명단</button>
                        <button onclick="switchTab('classes'); toggleMobileDrawer();" id="m-tab-classes" class="w-full text-left px-3.5 py-2.5 text-slate-600 hover:bg-slate-50 font-bold rounded-xl transition">좌석표 설정 및 인쇄</button>
                        <button onclick="switchTab('qr'); toggleMobileDrawer();" id="m-tab-qr" class="w-full text-left px-3.5 py-2.5 text-slate-600 hover:bg-slate-50 font-bold rounded-xl transition">로그인 QR</button>
                        <a href="/" target="_blank" class="block px-3.5 py-2.5 text-slate-800 hover:bg-slate-50 font-black rounded-xl mt-4 border-t pt-4 transition">좌석표 신청(학생용)</a>
                    </nav>
                </div>
                <div class="text-[11px] text-slate-400 text-center font-bold">김영편입 좌석배치 시스템</div>
            </div>
        </div>

        <aside class="w-64 bg-white border-r border-slate-200 flex-col hidden md:flex p-5 shadow-sm">
            <h1 class="text-lg font-bold text-blue-700 mb-6 tracking-tight">김영편입 좌석표 신청</h1>
            <nav class="space-y-1.5 flex-1">
                <button onclick="switchTab('students')" id="tab-btn-students" class="w-full text-left px-3.5 py-2.5 bg-blue-50 text-blue-700 font-bold rounded-xl transition">학생명단</button>
                <button onclick="switchTab('classes')" id="tab-btn-classes" class="w-full text-left px-3.5 py-2.5 text-slate-600 hover:bg-slate-50 font-bold rounded-xl transition">좌석표 설정 및 인쇄</button>
                <button onclick="switchTab('qr')" id="tab-btn-qr" class="w-full text-left px-3.5 py-2.5 text-slate-600 hover:bg-slate-50 font-bold rounded-xl transition">로그인 QR</button>
                <a href="/" target="_blank" class="block px-3.5 py-2.5 text-slate-800 hover:bg-slate-50 font-black rounded-xl mt-4 border-t pt-4 transition">좌석표 신청(학생용)</a>
            </nav>
            <div class="border-t pt-3 text-center">
                <div id="desktop-live-date" class="text-[11px] font-bold text-slate-400"></div>
                <div id="desktop-live-clock" class="text-sm font-black text-slate-700 font-mono mt-0.5"></div>
            </div>
        </aside>

        <main class="flex-1 overflow-y-auto p-4 md:p-8" id="admin-main-area">
            
            <!-- 탭 1: 학생명단 -->
            <div id="tab-students">
                <div class="flex flex-col md:flex-row justify-between items-start md:items-center gap-3 mb-4">
                    <div>
                        <div class="flex items-center gap-3">
                            <h2 class="text-xl md:text-2xl font-bold text-slate-900">학생명단</h2>
                            <button onclick="toggleClassFilterPanel()" id="btn-class-filter-toggle" class="bg-slate-100 hover:bg-slate-200 text-slate-700 border border-slate-300 px-3 py-1.5 rounded-lg text-xs font-bold transition flex items-center gap-1.5 shadow-xs">
                                📋 반별보기 <span id="filter-active-count" class="bg-blue-600 text-white px-1.5 py-0.5 rounded-full text-[10px] hidden font-bold">0</span>
                            </button>
                        </div>
                        <p id="summary-text" class="text-xs md:text-sm text-slate-500 mt-1 font-medium">불러오는 중...</p>
                    </div>
                    <div class="flex flex-wrap gap-2">
                        <button onclick="clearAllStudents()" class="bg-rose-50 hover:bg-rose-100 text-rose-600 border border-rose-200 px-3 py-2 rounded-lg text-xs md:text-sm font-bold transition shadow-xs">
                            ⚠️ 명단 초기화
                        </button>
                        <button onclick="openAddStudentModal()" class="bg-emerald-600 hover:bg-emerald-700 text-white px-3 py-2 rounded-lg text-xs md:text-sm font-bold shadow-sm transition">
                            ➕ 학생 수동 추가
                        </button>
                        <label class="cursor-pointer bg-blue-600 hover:bg-blue-700 text-white px-3 py-2 rounded-lg text-xs md:text-sm font-bold shadow-sm transition">
                            수강생 파일 업로드
                            <input type="file" id="upload-student-file" accept=".xls,.xlsx,.csv,.htm,.html" class="hidden" onchange="uploadStudentFile()">
                        </label>
                    </div>
                </div>

                <div id="class-filter-panel" class="hidden bg-white border border-slate-200 rounded-xl p-4 shadow-md mb-6">
                    <div class="flex justify-between items-center pb-2 border-b border-slate-100 mb-3">
                        <span class="text-xs font-bold text-slate-700">반 선택 필터</span>
                        <div class="flex gap-2 text-xs">
                            <button onclick="selectAllClassFilters(true)" class="text-blue-600 font-bold hover:underline">전체 선택</button>
                            <span class="text-slate-300">|</span>
                            <button onclick="selectAllClassFilters(false)" class="text-slate-500 font-bold hover:underline">전체 해제</button>
                        </div>
                    </div>
                    <div id="class-filter-checkboxes" class="grid grid-cols-2 sm:grid-cols-3 md:grid-cols-4 gap-2.5 max-h-48 overflow-y-auto"></div>
                </div>

                <div class="bg-white p-3 md:p-4 rounded-xl border border-slate-200 mb-4 md:mb-6 shadow-sm">
                    <input type="text" id="search-keyword" oninput="fetchStudents()" placeholder="아이디, 이름 또는 반명 검색..." class="w-full border rounded-lg px-3 py-2 text-sm outline-none focus:border-blue-500 font-medium">
                </div>
                <div class="bg-white rounded-xl border border-slate-200 overflow-x-auto shadow-sm">
                    <table class="w-full text-left text-sm min-w-[550px]">
                        <thead class="bg-slate-50 border-b text-xs font-bold text-slate-500 select-none">
                            <tr>
                                <th onclick="toggleSort('username')" class="py-3 px-4 w-40 cursor-pointer hover:bg-slate-100 transition">
                                    <div class="flex items-center gap-1">아이디 <span id="sort-icon-username" class="text-slate-400">↕</span></div>
                                </th>
                                <th onclick="toggleSort('name')" class="py-3 px-4 w-32 cursor-pointer hover:bg-slate-100 transition">
                                    <div class="flex items-center gap-1">이름 <span id="sort-icon-name" class="text-slate-400">↕</span></div>
                                </th>
                                <th onclick="toggleSort('class_name')" class="py-3 px-4 cursor-pointer hover:bg-slate-100 transition">
                                    <div class="flex items-center gap-1">수강 반 목록 <span id="sort-icon-class_name" class="text-slate-400">↕</span></div>
                                </th>
                                <th class="py-3 px-3 w-24 text-center">비밀번호</th>
                                <th class="py-3 px-3 w-20 text-center">관리</th>
                            </tr>
                        </thead>
                        <tbody id="student-tbody" class="divide-y divide-slate-100"></tbody>
                    </table>
                </div>
            </div>

            <!-- 탭 2: 좌석표 설정 및 인쇄 -->
            <div id="tab-classes" class="hidden">
                <div class="flex flex-col md:flex-row justify-between items-start md:items-center gap-3 mb-6">
                    <div>
                        <h2 class="text-xl md:text-2xl font-bold text-slate-900">좌석표 설정 및 인쇄</h2>
                        <p class="text-xs md:text-sm text-slate-500 mt-1 font-medium">강의실별 좌석수가 [301호 (48석)] 형태로 정확히 계산되어 표시됩니다.</p>
                    </div>
                    <label class="cursor-pointer bg-emerald-600 hover:bg-emerald-700 text-white px-4 py-2 rounded-lg text-sm font-bold shadow-sm transition">
                        강의실 좌석표 엑셀 업로드 (.xlsx)
                        <input type="file" id="upload-room-file" accept=".xlsx" class="hidden" onchange="uploadRoomFile()">
                    </label>
                </div>

                <div class="bg-white rounded-xl border border-slate-200 overflow-x-auto shadow-sm">
                    <table class="w-full text-left text-sm min-w-[760px]">
                        <thead class="bg-slate-50 border-b text-xs font-bold text-slate-500">
                            <tr>
                                <th class="py-3.5 px-4">반(강좌)명</th>
                                <th class="py-3.5 px-3 text-center w-16">총원</th>
                                <th class="py-3.5 px-3 text-center w-20">신청</th>
                                <th class="py-3.5 px-3 text-center w-20">미신청</th>
                                <th class="py-3.5 px-2 text-center w-16">행</th>
                                <th class="py-3.5 px-2 text-center w-16">열</th>
                                <th class="py-3.5 px-3 w-40">배정 강의실(좌석수)</th>
                                <th class="py-3.5 px-3 w-52 text-center">자동 리셋 주기 / 일시</th>
                                <th class="py-3.5 px-2 text-center w-20">저장</th>
                                <th class="py-3.5 px-2 text-center w-20">즉시리셋</th>
                                <th class="py-3.5 px-2 text-center w-16">인쇄</th>
                            </tr>
                        </thead>
                        <tbody id="class-config-tbody" class="divide-y divide-slate-100"></tbody>
                    </table>
                </div>
            </div>

            <!-- 탭 3: 로그인 QR -->
            <div id="tab-qr" class="hidden" onclick="toggleQrFullscreenHeader()">
                <div class="mb-4 flex flex-col sm:flex-row justify-between items-start sm:items-center gap-3">
                    <div>
                        <h2 class="text-xl md:text-2xl font-bold text-slate-900">로그인 QR</h2>
                        <p class="text-xs text-slate-500 mt-0.5 font-medium">화면을 가볍게 터치하면 상단 메뉴바가 토글됩니다.</p>
                    </div>
                    <button onclick="window.print()" class="bg-slate-800 text-white px-4 py-2 rounded-lg text-xs md:text-sm font-bold hover:bg-slate-900 transition shadow-sm">포스터 인쇄 / PDF 저장</button>
                </div>

                <div class="flex justify-center items-center py-4">
                    <div class="bg-white border-2 border-blue-600 rounded-3xl p-6 sm:p-8 max-w-sm w-full shadow-xl text-center">
                        <div class="bg-blue-50 border border-blue-200 rounded-xl p-3 mb-4">
                            <div id="qr-live-date" class="text-xs font-bold text-blue-900"></div>
                            <div id="qr-live-clock" class="text-xl font-black text-blue-600 font-mono mt-0.5">00:00:00</div>
                        </div>

                        <div class="text-lg font-extrabold text-slate-900 mb-4" id="qr-date-title">좌석표 신청</div>
                        
                        <div class="bg-slate-50 p-4 rounded-2xl inline-block border border-slate-200 mb-4 shadow-inner">
                            <div id="qrcode" class="flex justify-center"></div>
                        </div>

                        <p class="text-xs font-bold text-blue-600 mb-1">카메라로 스캔 ➔ 로그인 ➔ 좌석 선택</p>
                        <p class="text-[11px] text-slate-400 font-medium">초기 비밀번호는 1234 입니다.</p>
                    </div>
                </div>
            </div>
        </main>

        <!-- 신청 완료 학생 명단 모달 -->
        <div id="reserved-modal" class="fixed inset-0 bg-slate-900/50 hidden items-center justify-center z-50 p-4">
            <div class="bg-white rounded-2xl p-6 w-full max-w-md shadow-2xl flex flex-col max-h-[80vh]">
                <div class="flex justify-between items-center pb-3 border-b border-slate-100">
                    <div>
                        <h3 class="text-base font-extrabold text-slate-900" id="reserved-modal-title">좌석 신청자 명단</h3>
                        <p class="text-xs text-blue-600 font-bold mt-0.5" id="reserved-modal-count"></p>
                    </div>
                    <button onclick="closeReservedModal()" class="text-slate-400 hover:text-slate-600 font-bold text-xl">&times;</button>
                </div>

                <div class="overflow-y-auto py-3 space-y-2 flex-1 my-2" id="reserved-student-list"></div>

                <div class="pt-3 border-t border-slate-100 flex justify-end">
                    <button onclick="closeReservedModal()" class="px-4 py-2 bg-slate-800 text-white text-xs font-bold rounded-lg hover:bg-slate-900">닫기</button>
                </div>
            </div>
        </div>

        <!-- 미신청 학생 명단 모달 -->
        <div id="unreserved-modal" class="fixed inset-0 bg-slate-900/50 hidden items-center justify-center z-50 p-4">
            <div class="bg-white rounded-2xl p-6 w-full max-w-md shadow-2xl flex flex-col max-h-[80vh]">
                <div class="flex justify-between items-center pb-3 border-b border-slate-100">
                    <div>
                        <h3 class="text-base font-extrabold text-slate-900" id="unreserved-modal-title">미신청자 명단</h3>
                        <p class="text-xs text-rose-500 font-bold mt-0.5" id="unreserved-modal-count"></p>
                    </div>
                    <button onclick="closeUnreservedModal()" class="text-slate-400 hover:text-slate-600 font-bold text-xl">&times;</button>
                </div>

                <div class="overflow-y-auto py-3 space-y-2 flex-1 my-2" id="unreserved-student-list"></div>

                <div class="pt-3 border-t border-slate-100 flex justify-end">
                    <button onclick="closeUnreservedModal()" class="px-4 py-2 bg-slate-800 text-white text-xs font-bold rounded-lg hover:bg-slate-900">닫기</button>
                </div>
            </div>
        </div>

        <!-- 달력 & 주기 팝업 모달 -->
        <div id="datetime-picker-modal" class="fixed inset-0 bg-slate-900/50 hidden items-center justify-center z-50 p-4">
            <div class="bg-white rounded-2xl p-6 w-full max-w-md shadow-2xl flex flex-col">
                <div class="flex justify-between items-center pb-3 border-b border-slate-100 mb-4">
                    <h3 class="text-base font-extrabold text-slate-900" id="dt-modal-title">자동 리셋 일시 설정</h3>
                    <button onclick="closeDateTimePickerModal()" class="text-slate-400 hover:text-slate-600 font-bold text-xl">&times;</button>
                </div>

                <div class="space-y-4 mb-6">
                    <div>
                        <label class="block text-xs font-bold text-slate-600 mb-1">리셋 주기</label>
                        <select id="dt-modal-repeat-type" onchange="onResetRepeatTypeChange()" class="w-full border border-slate-300 rounded-lg p-2.5 text-sm font-bold text-slate-800 bg-white cursor-pointer outline-none focus:border-blue-500">
                            <option value="once">📅 특정 날짜 (1회 리셋)</option>
                            <option value="daily">🔄 매일 반복</option>
                            <option value="weekly">📆 매주 반복 (요일 선택)</option>
                            <option value="monthly">🗓️ 매월 반복 (날짜 선택)</option>
                        </select>
                    </div>

                    <div id="dt-modal-date-wrapper">
                        <label class="block text-xs font-bold text-slate-600 mb-1">달력에서 날짜 선택</label>
                        <input type="date" id="dt-modal-date" class="w-full border border-slate-300 rounded-lg p-2.5 text-sm font-bold text-slate-800 outline-none focus:border-blue-500 bg-slate-50 cursor-pointer">
                    </div>

                    <div id="dt-modal-weekday-wrapper" class="hidden">
                        <label class="block text-xs font-bold text-slate-600 mb-1">매주 반복할 요일</label>
                        <select id="dt-modal-weekday" class="w-full border border-slate-300 rounded-lg p-2.5 text-sm font-bold text-slate-800 bg-white cursor-pointer">
                            <option value="0">매주 월요일</option>
                            <option value="1">매주 화요일</option>
                            <option value="2">매주 수요일</option>
                            <option value="3">매주 목요일</option>
                            <option value="4">매주 금요일</option>
                            <option value="5">매주 토요일</option>
                            <option value="6">매주 일요일</option>
                        </select>
                    </div>

                    <div id="dt-modal-day-wrapper" class="hidden">
                        <label class="block text-xs font-bold text-slate-600 mb-1">매월 반복할 일자</label>
                        <select id="dt-modal-day" class="w-full border border-slate-300 rounded-lg p-2.5 text-sm font-bold text-slate-800 bg-white cursor-pointer"></select>
                    </div>

                    <div>
                        <label class="block text-xs font-bold text-slate-600 mb-1">리셋 시간</label>
                        <div class="grid grid-cols-3 gap-2">
                            <select id="dt-modal-ampm" class="border border-slate-300 rounded-lg p-2.5 text-sm font-bold text-slate-800 bg-white cursor-pointer">
                                <option value="AM">오전</option>
                                <option value="PM" selected>오후</option>
                            </select>
                            <select id="dt-modal-hour" class="border border-slate-300 rounded-lg p-2.5 text-sm font-bold text-slate-800 bg-white cursor-pointer"></select>
                            <select id="dt-modal-minute" class="border border-slate-300 rounded-lg p-2.5 text-sm font-bold text-slate-800 bg-white cursor-pointer"></select>
                        </div>
                    </div>
                </div>

                <div class="flex justify-between items-center pt-3 border-t border-slate-100">
                    <button onclick="clearDateTimePickerModal()" class="px-3.5 py-2 bg-rose-50 text-rose-600 text-xs font-bold rounded-lg hover:bg-rose-100 transition">리셋 해제</button>
                    <div class="flex gap-2">
                        <button onclick="closeDateTimePickerModal()" class="px-3.5 py-2 bg-slate-200 text-slate-700 text-xs font-bold rounded-lg hover:bg-slate-300 transition">취소</button>
                        <button onclick="confirmDateTimePickerModal()" class="px-5 py-2 bg-blue-600 text-white text-xs font-bold rounded-lg hover:bg-blue-700 shadow-sm transition">설정 적용</button>
                    </div>
                </div>
            </div>
        </div>

        <!-- 학생 수동 추가 모달 -->
        <div id="add-student-modal" class="fixed inset-0 bg-slate-900/50 hidden items-center justify-center z-50 p-4">
            <div class="bg-white rounded-2xl p-6 w-full max-w-md shadow-2xl flex flex-col">
                <div class="flex justify-between items-center pb-3 border-b border-slate-100 mb-4">
                    <h3 class="text-base font-bold text-slate-900">➕ 학생 수동 추가</h3>
                    <button onclick="closeAddStudentModal()" class="text-slate-400 hover:text-slate-600 font-bold text-xl">&times;</button>
                </div>

                <div class="space-y-4 mb-6">
                    <div>
                        <label class="block text-xs font-bold text-slate-600 mb-1">아이디</label>
                        <input type="text" id="add-username" placeholder="예: kimyoung123" class="w-full border border-slate-300 rounded-lg p-2.5 text-sm font-bold outline-none focus:border-blue-500">
                    </div>

                    <div>
                        <label class="block text-xs font-bold text-slate-600 mb-1">이름</label>
                        <input type="text" id="add-name" placeholder="예: 홍길동" class="w-full border border-slate-300 rounded-lg p-2.5 text-sm font-bold outline-none focus:border-blue-500">
                    </div>

                    <div>
                        <label class="block text-xs font-bold text-slate-600 mb-1">수강 반 선택 (복수 선택 가능)</label>
                        <div id="add-class-options" class="border border-slate-200 rounded-lg p-3 max-h-40 overflow-y-auto space-y-1.5 bg-slate-50"></div>
                    </div>
                </div>

                <div class="flex justify-end gap-2 pt-3 border-t border-slate-100">
                    <button onclick="closeAddStudentModal()" class="px-4 py-2 bg-slate-200 text-slate-700 text-xs font-bold rounded-lg hover:bg-slate-300">취소</button>
                    <button onclick="submitAddStudent()" class="px-5 py-2 bg-emerald-600 text-white text-xs font-bold rounded-lg hover:bg-emerald-700 shadow-sm">추가 등록</button>
                </div>
            </div>
        </div>

        <script>
            let availableRooms = [];
            let currentSortColumn = "username";
            let currentSortOrder = "asc";
            let selectedClassFilters = [];
            let targetConfigIdx = -1;

            const weekdayNames = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"];

            function updateAdminClocks() {
                const now = new Date();
                const days = ['일', '월', '화', '수', '목', '금', '토'];
                const year = now.getFullYear();
                const month = String(now.getMonth() + 1).padStart(2, '0');
                const date = String(now.getDate()).padStart(2, '0');
                const dayName = days[now.getDay()];
                
                const hours = String(now.getHours()).padStart(2, '0');
                const minutes = String(now.getMinutes()).padStart(2, '0');
                const seconds = String(now.getSeconds()).padStart(2, '0');

                const dateStr = `${year}년 ${month}월 ${date}일 (${dayName})`;
                const clockStr = `${hours}:${minutes}:${seconds}`;

                const headerClock = document.getElementById("header-clock");
                if (headerClock) headerClock.innerText = clockStr;

                const dDate = document.getElementById("desktop-live-date");
                const dClock = document.getElementById("desktop-live-clock");
                if (dDate) dDate.innerText = dateStr;
                if (dClock) dClock.innerText = clockStr;

                const qrDate = document.getElementById("qr-live-date");
                const qrClock = document.getElementById("qr-live-clock");
                if (qrDate) qrDate.innerText = dateStr;
                if (qrClock) qrClock.innerText = clockStr;
            }
            setInterval(updateAdminClocks, 1000);
            updateAdminClocks();

            function toggleMobileDrawer() {
                const drawer = document.getElementById("mobile-drawer");
                drawer.classList.toggle("hidden");
            }

            function toggleQrFullscreenHeader() {
                const header = document.getElementById("mobile-header");
                if(window.innerWidth < 768) {
                    header.classList.toggle("hidden");
                }
            }

            function switchTab(tab) {
                document.getElementById("tab-students").classList.add("hidden");
                document.getElementById("tab-classes").classList.add("hidden");
                document.getElementById("tab-qr").classList.add("hidden");

                ["tab-btn-students", "tab-btn-classes", "tab-btn-qr", "m-tab-students", "m-tab-classes", "m-tab-qr"].forEach(id => {
                    const el = document.getElementById(id);
                    if(el) {
                        el.className = el.className.replace("bg-blue-50 text-blue-700", "text-slate-600 hover:bg-slate-50");
                    }
                });

                if (tab === 'students') {
                    document.getElementById("tab-students").classList.remove("hidden");
                    document.getElementById("tab-btn-students").className = "w-full text-left px-3.5 py-2.5 bg-blue-50 text-blue-700 font-bold rounded-xl transition";
                    document.getElementById("m-tab-students").className = "w-full text-left px-3.5 py-2.5 bg-blue-50 text-blue-700 font-bold rounded-xl transition";
                    fetchStudents();
                } else if (tab === 'classes') {
                    document.getElementById("tab-classes").classList.remove("hidden");
                    document.getElementById("tab-btn-classes").className = "w-full text-left px-3.5 py-2.5 bg-blue-50 text-blue-700 font-bold rounded-xl transition";
                    document.getElementById("m-tab-classes").className = "w-full text-left px-3.5 py-2.5 bg-blue-50 text-blue-700 font-bold rounded-xl transition";
                    fetchClassConfigs();
                } else if (tab === 'qr') {
                    document.getElementById("tab-qr").classList.remove("hidden");
                    document.getElementById("tab-btn-qr").className = "w-full text-left px-3.5 py-2.5 bg-blue-50 text-blue-700 font-bold rounded-xl transition";
                    document.getElementById("m-tab-qr").className = "w-full text-left px-3.5 py-2.5 bg-blue-50 text-blue-700 font-bold rounded-xl transition";
                    generateTodayQR();
                }
            }

            async function clearAllStudents() {
                if(!confirm("⚠️ 경고: 등록된 모든 학생 정보 및 신청 내역이 삭제됩니다.\\n정말 전체 명단을 초기화하시겠습니까?")) return;
                const res = await fetch("/api/admin/students/clear-all", { method: "POST" });
                const data = await res.json();
                alert(data.message);
                fetchStudents();
            }

            function toggleSort(column) {
                if (currentSortColumn === column) {
                    currentSortOrder = (currentSortOrder === "asc") ? "desc" : "asc";
                } else {
                    currentSortColumn = column;
                    currentSortOrder = "asc";
                }
                updateSortIcons();
                fetchStudents();
            }

            function updateSortIcons() {
                ["username", "name", "class_name"].forEach(col => {
                    const iconSpan = document.getElementById(`sort-icon-${col}`);
                    if (iconSpan) {
                        if (currentSortColumn === col) {
                            iconSpan.innerText = currentSortOrder === "asc" ? "▲" : "▼";
                            iconSpan.className = "text-blue-600 font-bold";
                        } else {
                            iconSpan.innerText = "↕";
                            iconSpan.className = "text-slate-400 font-normal";
                        }
                    }
                });
            }

            async function toggleClassFilterPanel() {
                const panel = document.getElementById("class-filter-panel");
                if (panel.classList.contains("hidden")) {
                    panel.classList.remove("hidden");
                    await loadClassFilterList();
                } else {
                    panel.classList.add("hidden");
                }
            }

            async function loadClassFilterList() {
                const res = await fetch("/api/admin/classes");
                const data = await res.json();
                const container = document.getElementById("class-filter-checkboxes");
                container.innerHTML = "";
                
                if (data.classes.length === 0) {
                    container.innerHTML = `<div class="text-xs text-slate-400 col-span-full py-2 text-center font-bold">등록된 반이 없습니다.</div>`;
                    return;
                }
                
                data.classes.forEach(c => {
                    const isChecked = selectedClassFilters.includes(c.class_name);
                    container.innerHTML += `
                        <label class="flex items-center gap-2 cursor-pointer p-1.5 hover:bg-slate-50 rounded-lg border border-slate-100 transition">
                            <input type="checkbox" value="${c.class_name}" ${isChecked ? 'checked' : ''} onchange="onClassFilterChange()" class="class-filter-cb w-4 h-4 text-blue-600 rounded">
                            <span class="text-xs font-bold text-slate-800 truncate" title="${c.class_name}">${c.class_name}</span>
                        </label>
                    `;
                });
            }

            function onClassFilterChange() {
                const checkboxes = document.querySelectorAll('.class-filter-cb:checked');
                selectedClassFilters = Array.from(checkboxes).map(cb => cb.value);
                const activeBadge = document.getElementById("filter-active-count");
                if (selectedClassFilters.length > 0) {
                    activeBadge.innerText = selectedClassFilters.length;
                    activeBadge.classList.remove("hidden");
                } else {
                    activeBadge.classList.add("hidden");
                }
                fetchStudents();
            }

            function selectAllClassFilters(select) {
                document.querySelectorAll('.class-filter-cb').forEach(cb => cb.checked = select);
                onClassFilterChange();
            }

            async function fetchStudents() {
                const keyword = document.getElementById("search-keyword").value;
                const classFilterStr = selectedClassFilters.join(",");
                const res = await fetch(`/api/admin/students?keyword=${encodeURIComponent(keyword)}&sort_by=${currentSortColumn}&order=${currentSortOrder}&class_filter=${encodeURIComponent(classFilterStr)}`);
                const data = await res.json();
                
                if (selectedClassFilters.length > 0) {
                    document.getElementById("summary-text").innerText = `총 등록 학생: ${data.total}명 | 필터링: ${data.filtered_total}명`;
                } else {
                    document.getElementById("summary-text").innerText = `총 등록 학생: ${data.total}명`;
                }

                const tbody = document.getElementById("student-tbody");
                tbody.innerHTML = "";
                data.students.forEach(s => {
                    tbody.innerHTML += `
                        <tr class="hover:bg-slate-50 font-medium text-xs md:text-sm">
                            <td class="py-2.5 px-4 font-mono text-slate-700 font-bold">${s.username}</td>
                            <td class="py-2.5 px-4 font-bold text-slate-900">${s.name}</td>
                            <td class="py-2.5 px-4 text-slate-600 font-medium">${s.class_name}</td>
                            <td class="py-2.5 px-3 text-center">
                                <button onclick="resetPassword('${s.username}')" class="text-xs bg-slate-100 hover:bg-slate-200 px-2 py-1 rounded border font-bold">초기화</button>
                            </td>
                            <td class="py-2.5 px-3 text-center">
                                <button onclick="deleteStudent('${s.username}')" class="text-xs text-red-500 hover:text-red-700 font-bold">삭제</button>
                            </td>
                        </tr>
                    `;
                });
            }

            async function openAddStudentModal() {
                document.getElementById("add-username").value = "";
                document.getElementById("add-name").value = "";
                
                const res = await fetch("/api/admin/classes");
                const data = await res.json();
                const classContainer = document.getElementById("add-class-options");
                classContainer.innerHTML = "";
                
                if (data.classes.length === 0) {
                    classContainer.innerHTML = `<div class="text-xs text-slate-400 py-2 text-center font-bold">등록된 반이 없습니다.</div>`;
                } else {
                    data.classes.forEach(c => {
                        classContainer.innerHTML += `
                            <label class="flex items-center gap-2 cursor-pointer p-1 hover:bg-white rounded transition">
                                <input type="checkbox" name="add-student-class" value="${c.class_name}" class="w-4 h-4 text-blue-600 rounded">
                                <span class="text-xs font-bold text-slate-800">${c.class_name}</span>
                            </label>
                        `;
                    });
                }
                document.getElementById("add-student-modal").classList.replace("hidden", "flex");
            }

            function closeAddStudentModal() { document.getElementById("add-student-modal").classList.replace("flex", "hidden"); }

            async function submitAddStudent() {
                const username = document.getElementById("add-username").value.trim();
                const name = document.getElementById("add-name").value.trim();
                const checkboxes = document.querySelectorAll('input[name="add-student-class"]:checked');
                const selectedClasses = Array.from(checkboxes).map(cb => cb.value);

                if (!username || !name) { alert("아이디와 이름을 입력해 주세요."); return; }
                if (selectedClasses.length === 0) { alert("수강 반을 최소 하나 선택해 주세요."); return; }

                const res = await fetch("/api/admin/students/add-manual", {
                    method: "POST",
                    headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                    body: `username=${encodeURIComponent(username)}&name=${encodeURIComponent(name)}&class_names=${encodeURIComponent(selectedClasses.join(', '))}`
                });
                const data = await res.json();
                alert(data.message);
                if (data.success) {
                    closeAddStudentModal();
                    fetchStudents();
                }
            }

            function formatResetDisplay(dtStr) {
                if (!dtStr) return "일시 미설정";
                
                if (dtStr.startsWith("DAILY:")) {
                    const tStr = dtStr.replace("DAILY:", "");
                    let [h, m] = tStr.split(":");
                    let hNum = parseInt(h, 10);
                    const ampm = hNum >= 12 ? "오후" : "오전";
                    if (hNum > 12) hNum -= 12;
                    if (hNum === 0) hNum = 12;
                    const hDisplay = hNum < 10 ? '0' + hNum : '' + hNum;
                    return `매일 ${ampm} ${hDisplay}:${m}`;
                }

                if (dtStr.startsWith("WEEKLY:")) {
                    const parts = dtStr.replace("WEEKLY:", "").split(":");
                    const wdayName = weekdayNames[parseInt(parts[0], 10)] || "";
                    let hNum = parseInt(parts[1], 10);
                    const ampm = hNum >= 12 ? "오후" : "오전";
                    if (hNum > 12) hNum -= 12;
                    if (hNum === 0) hNum = 12;
                    const hDisplay = hNum < 10 ? '0' + hNum : '' + hNum;
                    return `매주(${wdayName}) ${ampm} ${hDisplay}:${parts[2]}`;
                }

                if (dtStr.startsWith("MONTHLY:")) {
                    const parts = dtStr.replace("MONTHLY:", "").split(":");
                    let hNum = parseInt(parts[1], 10);
                    const ampm = hNum >= 12 ? "오후" : "오전";
                    if (hNum > 12) hNum -= 12;
                    if (hNum === 0) hNum = 12;
                    const hDisplay = hNum < 10 ? '0' + hNum : '' + hNum;
                    return `매월(${parts[0]}일) ${ampm} ${hDisplay}:${parts[2]}`;
                }

                if (dtStr.includes("T")) {
                    const [dStr, tStr] = dtStr.split("T");
                    let [h, m] = tStr.split(":");
                    let hNum = parseInt(h, 10);
                    const ampm = hNum >= 12 ? "오후" : "오전";
                    if (hNum > 12) hNum -= 12;
                    if (hNum === 0) hNum = 12;
                    const hDisplay = hNum < 10 ? '0' + hNum : '' + hNum;
                    return `${dStr} ${ampm} ${hDisplay}:${m}`;
                }

                return "일시 미설정";
            }

            async function fetchClassConfigs() {
                const res = await fetch("/api/admin/classes");
                const data = await res.json();
                availableRooms = data.rooms;

                const tbody = document.getElementById("class-config-tbody");
                tbody.innerHTML = "";

                data.classes.forEach((c, idx) => {
                    let options = availableRooms.map(r => {
                        const isSel = r.room_name === c.room_name ? 'selected' : '';
                        return `<option value="${r.room_name}" ${isSel}>${r.room_name} (${r.total_seats}석)</option>`;
                    }).join("");
                    if (!options) options = `<option value="${c.room_name}">${c.room_name}</option>`;

                    const reservedBadge = c.reserved_count > 0 
                        ? `<button onclick="showReserved('${c.class_name}')" class="px-2 py-0.5 bg-blue-50 text-blue-700 border border-blue-200 rounded-full text-xs font-extrabold hover:bg-blue-100 transition shadow-xs">${c.reserved_count}명 🔍</button>`
                        : `<span class="px-2 py-0.5 bg-slate-100 text-slate-400 rounded-full text-xs font-bold">0명</span>`;

                    const unreservedBadge = c.unreserved_count > 0 
                        ? `<button onclick="showUnreserved('${c.class_name}')" class="px-2 py-0.5 bg-rose-50 text-rose-600 border border-rose-200 rounded-full text-xs font-extrabold hover:bg-rose-100 transition shadow-xs">${c.unreserved_count}명 🔍</button>`
                        : `<span class="px-2 py-0.5 bg-slate-100 text-slate-400 rounded-full text-xs font-bold">0명</span>`;

                    const displayDtText = formatResetDisplay(c.reset_datetime);

                    tbody.innerHTML += `
                        <tr class="hover:bg-slate-50 font-medium text-xs md:text-sm">
                            <td class="py-3 px-4 font-bold text-slate-800">${c.class_name}</td>
                            <td class="py-3 px-1 text-center font-bold text-slate-700">${c.total_students}명</td>
                            <td class="py-3 px-1 text-center">${reservedBadge}</td>
                            <td class="py-3 px-1 text-center">${unreservedBadge}</td>
                            <td class="py-3 px-1 text-center">
                                <input type="number" id="rows-${idx}" value="${c.rows_count}" min="1" max="30" class="border rounded px-1 py-1 text-xs w-12 text-center font-bold text-blue-700">
                            </td>
                            <td class="py-3 px-1 text-center">
                                <input type="number" id="cols-${idx}" value="${c.cols_count}" min="1" max="30" class="border rounded px-1 py-1 text-xs w-12 text-center font-bold text-blue-700">
                            </td>
                            <td class="py-3 px-2">
                                <select id="room-${idx}" class="border rounded px-1.5 py-1 text-xs w-full bg-white font-bold cursor-pointer" onchange="onRoomSelectChanged(${idx})">
                                    ${options}
                                </select>
                            </td>
                            <td class="py-3 px-2 text-center">
                                <input type="hidden" id="reset-dt-${idx}" value="${c.reset_datetime || ''}">
                                <input type="text" id="reset-dt-display-${idx}" value="${displayDtText}" readonly onclick="openDateTimePickerModal(${idx}, '${c.class_name}')" class="border border-slate-300 rounded px-2 py-1 text-xs font-bold text-slate-700 bg-white hover:bg-blue-50/50 cursor-pointer w-48 text-center transition shadow-xs" title="클릭하여 달력 및 주기 선택">
                            </td>
                            <td class="py-3 px-1 text-center">
                                <button onclick="saveClassConfig('${c.class_name}', ${idx})" id="save-btn-${idx}" class="text-xs bg-blue-600 text-white px-3 py-1 rounded-lg hover:bg-blue-700 font-bold shadow-xs transition">저장</button>
                            </td>
                            <td class="py-3 px-1 text-center">
                                <button onclick="resetClassNow('${c.class_name}')" class="text-xs bg-rose-500 text-white px-2 py-1 rounded-lg hover:bg-rose-600 font-bold shadow-xs">초기화</button>
                            </td>
                            <td class="py-3 px-1 text-center">
                                <a href="/admin/print-seating-chart?class_name=${encodeURIComponent(c.class_name)}" target="_blank" class="inline-block text-xs bg-emerald-600 hover:bg-emerald-700 text-white font-bold px-2 py-1 rounded-lg shadow-xs">
                                    인쇄
                                </a>
                            </td>
                        </tr>
                    `;
                });
            }

            function onRoomSelectChanged(idx) {
                const selRoomName = document.getElementById(`room-${idx}`).value;
                const rData = availableRooms.find(r => r.room_name === selRoomName);
                if (rData) {
                    document.getElementById(`rows-${idx}`).value = rData.rows_count;
                    document.getElementById(`cols-${idx}`).value = rData.cols_count;
                }
            }

            async function showReserved(className) {
                const res = await fetch(`/api/admin/reserved-students?class_name=${encodeURIComponent(className)}`);
                const data = await res.json();

                document.getElementById("reserved-modal-title").innerText = `[${className}] 좌석 신청 완료 학생`;
                document.getElementById("reserved-modal-count").innerText = `총 ${data.reserved.length}명이 신청을 완료했습니다.`;

                const listContainer = document.getElementById("reserved-student-list");
                listContainer.innerHTML = "";

                if (data.reserved.length === 0) {
                    listContainer.innerHTML = `<div class="text-center py-6 text-slate-400 text-sm font-bold">아직 좌석을 신청한 학생이 없습니다.</div>`;
                } else {
                    data.reserved.forEach(s => {
                        listContainer.innerHTML += `
                            <div class="flex justify-between items-center p-3 bg-blue-50/50 border border-blue-200 rounded-xl text-sm">
                                <div class="flex items-center gap-2">
                                    <span class="font-extrabold text-blue-900">${s.name}</span>
                                    <span class="font-mono text-xs text-slate-500 font-bold">(${s.username})</span>
                                </div>
                                <span class="bg-blue-600 text-white text-xs font-black px-2.5 py-1 rounded-lg shadow-xs">${s.seat_id}번 좌석</span>
                            </div>
                        `;
                    });
                }
                document.getElementById("reserved-modal").classList.replace("hidden", "flex");
            }

            function closeReservedModal() { document.getElementById("reserved-modal").classList.replace("flex", "hidden"); }

            async function showUnreserved(className) {
                const res = await fetch(`/api/admin/unreserved-students?class_name=${encodeURIComponent(className)}`);
                const data = await res.json();

                document.getElementById("unreserved-modal-title").innerText = `[${className}] 미신청 학생`;
                document.getElementById("unreserved-modal-count").innerText = `총 ${data.unreserved.length}명이 좌석을 선택하지 않았습니다.`;

                const listContainer = document.getElementById("unreserved-student-list");
                listContainer.innerHTML = "";

                if (data.unreserved.length === 0) {
                    listContainer.innerHTML = `<div class="text-center py-6 text-emerald-600 text-sm font-bold">모든 학생이 좌석 신청을 완료했습니다!</div>`;
                } else {
                    data.unreserved.forEach(s => {
                        listContainer.innerHTML += `
                            <div class="flex justify-between items-center p-2.5 bg-slate-50 border border-slate-200 rounded-xl text-sm">
                                <span class="font-bold text-slate-800">${s.name}</span>
                                <span class="font-mono text-xs text-slate-500 font-bold">${s.username}</span>
                            </div>
                        `;
                    });
                }
                document.getElementById("unreserved-modal").classList.replace("hidden", "flex");
            }

            function closeUnreservedModal() { document.getElementById("unreserved-modal").classList.replace("flex", "hidden"); }

            function initDateTimeModalOptions() {
                const hourSelect = document.getElementById("dt-modal-hour");
                hourSelect.innerHTML = "";
                for(let h=1; h<=12; h++) {
                    const val = h < 10 ? '0' + h : '' + h;
                    hourSelect.innerHTML += `<option value="${val}">${val}시</option>`;
                }
                
                const minSelect = document.getElementById("dt-modal-minute");
                minSelect.innerHTML = "";
                for(let m=0; m<60; m+=5) {
                    const val = m < 10 ? '0' + m : '' + m;
                    minSelect.innerHTML += `<option value="${val}">${val}분</option>`;
                }

                const daySelect = document.getElementById("dt-modal-day");
                daySelect.innerHTML = "";
                for(let d=1; d<=31; d++) {
                    daySelect.innerHTML += `<option value="${d}">매월 ${d}일</option>`;
                }
            }

            function onResetRepeatTypeChange() {
                const type = document.getElementById("dt-modal-repeat-type").value;
                const dateWrap = document.getElementById("dt-modal-date-wrapper");
                const weekWrap = document.getElementById("dt-modal-weekday-wrapper");
                const monthWrap = document.getElementById("dt-modal-day-wrapper");

                dateWrap.classList.add("hidden");
                weekWrap.classList.add("hidden");
                monthWrap.classList.add("hidden");

                if (type === "once") dateWrap.classList.remove("hidden");
                else if (type === "weekly") weekWrap.classList.remove("hidden");
                else if (type === "monthly") monthWrap.classList.remove("hidden");
            }

            function openDateTimePickerModal(idx, className) {
                targetConfigIdx = idx;
                initDateTimeModalOptions();

                document.getElementById("dt-modal-title").innerText = `[${className}] 자동 리셋 일시 설정`;

                const currentVal = document.getElementById(`reset-dt-${idx}`).value;
                const repeatSelect = document.getElementById("dt-modal-repeat-type");
                const dateInput = document.getElementById("dt-modal-date");
                const ampmSelect = document.getElementById("dt-modal-ampm");
                const hourSelect = document.getElementById("dt-modal-hour");
                const minSelect = document.getElementById("dt-modal-minute");
                const weekdaySelect = document.getElementById("dt-modal-weekday");
                const daySelect = document.getElementById("dt-modal-day");

                if (currentVal && currentVal.startsWith("DAILY:")) {
                    repeatSelect.value = "daily";
                    const tStr = currentVal.replace("DAILY:", "");
                    let [h, m] = tStr.split(":").map(Number);
                    if (h >= 12) { ampmSelect.value = "PM"; if (h > 12) h -= 12; }
                    else { ampmSelect.value = "AM"; if (h === 0) h = 12; }
                    hourSelect.value = h < 10 ? '0' + h : '' + h;
                    const mRounded = Math.floor(m / 5) * 5;
                    minSelect.value = mRounded < 10 ? '0' + mRounded : '' + mRounded;
                } else if (currentVal && currentVal.startsWith("WEEKLY:")) {
                    repeatSelect.value = "weekly";
                    const parts = currentVal.replace("WEEKLY:", "").split(":");
                    weekdaySelect.value = parts[0];
                    let h = parseInt(parts[1], 10);
                    let m = parseInt(parts[2], 10);
                    if (h >= 12) { ampmSelect.value = "PM"; if (h > 12) h -= 12; }
                    else { ampmSelect.value = "AM"; if (h === 0) h = 12; }
                    hourSelect.value = h < 10 ? '0' + h : '' + h;
                    const mRounded = Math.floor(m / 5) * 5;
                    minSelect.value = mRounded < 10 ? '0' + mRounded : '' + mRounded;
                } else if (currentVal && currentVal.startsWith("MONTHLY:")) {
                    repeatSelect.value = "monthly";
                    const parts = currentVal.replace("MONTHLY:", "").split(":");
                    daySelect.value = parts[0];
                    let h = parseInt(parts[1], 10);
                    let m = parseInt(parts[2], 10);
                    if (h >= 12) { ampmSelect.value = "PM"; if (h > 12) h -= 12; }
                    else { ampmSelect.value = "AM"; if (h === 0) h = 12; }
                    hourSelect.value = h < 10 ? '0' + h : '' + h;
                    const mRounded = Math.floor(m / 5) * 5;
                    minSelect.value = mRounded < 10 ? '0' + mRounded : '' + mRounded;
                } else if (currentVal && currentVal.includes("T")) {
                    repeatSelect.value = "once";
                    const [dStr, tStr] = currentVal.split("T");
                    dateInput.value = dStr;
                    let [h, m] = tStr.split(":").map(Number);
                    if (h >= 12) { ampmSelect.value = "PM"; if (h > 12) h -= 12; }
                    else { ampmSelect.value = "AM"; if (h === 0) h = 12; }
                    hourSelect.value = h < 10 ? '0' + h : '' + h;
                    const mRounded = Math.floor(m / 5) * 5;
                    minSelect.value = mRounded < 10 ? '0' + mRounded : '' + mRounded;
                } else {
                    repeatSelect.value = "once";
                    const now = new Date();
                    const yyyy = now.getFullYear();
                    const mm = String(now.getMonth() + 1).padStart(2, '0');
                    const dd = String(now.getDate()).padStart(2, '0');
                    dateInput.value = `${yyyy}-${mm}-${dd}`;
                    ampmSelect.value = "PM";
                    hourSelect.value = "12";
                    minSelect.value = "00";
                }

                onResetRepeatTypeChange();
                document.getElementById("datetime-picker-modal").classList.replace("hidden", "flex");
            }

            function closeDateTimePickerModal() {
                document.getElementById("datetime-picker-modal").classList.replace("flex", "hidden");
            }

            function clearDateTimePickerModal() {
                if (targetConfigIdx >= 0) {
                    document.getElementById(`reset-dt-${targetConfigIdx}`).value = "";
                    document.getElementById(`reset-dt-display-${targetConfigIdx}`).value = "일시 미설정";
                }
                closeDateTimePickerModal();
            }

            function confirmDateTimePickerModal() {
                if (targetConfigIdx < 0) return;

                const repeatType = document.getElementById("dt-modal-repeat-type").value;
                const ampm = document.getElementById("dt-modal-ampm").value;
                let h = parseInt(document.getElementById("dt-modal-hour").value, 10);
                const m = document.getElementById("dt-modal-minute").value;

                if (ampm === "PM" && h < 12) h += 12;
                if (ampm === "AM" && h === 12) h = 0;

                const hStr = h < 10 ? '0' + h : '' + h;
                const displayAmpm = ampm === "PM" ? "오후" : "오전";
                const displayHour = document.getElementById("dt-modal-hour").value;

                let saveVal = "";
                let displayVal = "";

                if (repeatType === "daily") {
                    saveVal = `DAILY:${hStr}:${m}`;
                    displayVal = `매일 ${displayAmpm} ${displayHour}:${m}`;
                } else if (repeatType === "weekly") {
                    const wVal = document.getElementById("dt-modal-weekday").value;
                    const wName = weekdayNames[parseInt(wVal, 10)];
                    saveVal = `WEEKLY:${wVal}:${hStr}:${m}`;
                    displayVal = `매주(${wName}) ${displayAmpm} ${displayHour}:${m}`;
                } else if (repeatType === "monthly") {
                    const dVal = document.getElementById("dt-modal-day").value;
                    saveVal = `MONTHLY:${dVal}:${hStr}:${m}`;
                    displayVal = `매월(${dVal}일) ${displayAmpm} ${displayHour}:${m}`;
                } else {
                    const dateVal = document.getElementById("dt-modal-date").value;
                    if (!dateVal) { alert("달력에서 날짜를 선택해주세요."); return; }
                    saveVal = `${dateVal}T${hStr}:${m}`;
                    displayVal = `${dateVal} ${displayAmpm} ${displayHour}:${m}`;
                }

                document.getElementById(`reset-dt-${targetConfigIdx}`).value = saveVal;
                document.getElementById(`reset-dt-display-${targetConfigIdx}`).value = displayVal;

                closeDateTimePickerModal();
            }

            async function saveClassConfig(className, idx) {
                const saveBtn = document.getElementById(`save-btn-${idx}`);
                const origText = saveBtn.innerText;
                saveBtn.disabled = true;
                saveBtn.innerText = "저장중...";

                const room = document.getElementById(`room-${idx}`).value;
                const rows = document.getElementById(`rows-${idx}`).value || 0;
                const cols = document.getElementById(`cols-${idx}`).value || 0;
                const resetDt = document.getElementById(`reset-dt-${idx}`).value || "";

                try {
                    const res = await fetch("/api/admin/classes/update", {
                        method: "POST",
                        headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                        body: `class_name=${encodeURIComponent(className)}&room_name=${encodeURIComponent(room)}&rows_count=${rows}&cols_count=${cols}&reset_datetime=${encodeURIComponent(resetDt)}`
                    });
                    const result = await res.json();

                    if (result.success) {
                        saveBtn.innerText = "✓ 완료";
                        saveBtn.className = "text-xs bg-emerald-600 text-white px-3 py-1 rounded-lg font-bold shadow-xs transition";
                        setTimeout(() => {
                            saveBtn.innerText = origText;
                            saveBtn.className = "text-xs bg-blue-600 text-white px-3 py-1 rounded-lg hover:bg-blue-700 font-bold shadow-xs transition";
                            saveBtn.disabled = false;
                        }, 1200);
                    } else {
                        alert(result.message);
                        saveBtn.innerText = origText;
                        saveBtn.disabled = false;
                    }
                } catch (e) {
                    alert("저장 중 오류가 발생했습니다.");
                    saveBtn.innerText = origText;
                    saveBtn.disabled = false;
                }
            }

            async function resetClassNow(className) {
                if (!confirm(`[${className}] 반의 좌석 신청 내역을 즉시 초기화하시겠습니까?`)) return;
                const res = await fetch("/api/admin/classes/reset-now", {
                    method: "POST",
                    headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                    body: `class_name=${encodeURIComponent(className)}`
                });
                const result = await res.json();
                alert(result.message);
                fetchClassConfigs();
            }

            async function uploadStudentFile() {
                const fileInput = document.getElementById("upload-student-file");
                if (!fileInput.files[0]) return;
                const formData = new FormData();
                formData.append("file", fileInput.files[0]);

                document.getElementById("summary-text").innerText = "명단 업로드 중...";
                const res = await fetch("/api/admin/students/upload", { method: "POST", body: formData });
                const result = await res.json();
                alert(result.message);
                fileInput.value = "";
                fetchStudents();
            }

            async function uploadRoomFile() {
                const fileInput = document.getElementById("upload-room-file");
                if (!fileInput.files[0]) return;
                const formData = new FormData();
                formData.append("file", fileInput.files[0]);

                const res = await fetch("/api/admin/rooms/upload", { method: "POST", body: formData });
                const result = await res.json();
                alert(result.message);
                fileInput.value = "";
                fetchClassConfigs();
            }

            async function generateTodayQR() {
                const currentHost = window.location.origin;
                const targetUrl = `${currentHost}/`;
                
                const qrContainer = document.getElementById("qrcode");
                qrContainer.innerHTML = "";
                new QRCode(qrContainer, {
                    text: targetUrl,
                    width: 200,
                    height: 200,
                    colorDark : "#0f172a",
                    colorLight : "#ffffff",
                    correctLevel : QRCode.CorrectLevel.H
                });
            }

            async function resetPassword(username) {
                if (!confirm(`${username} 학생의 비밀번호를 '1234'로 초기화하시겠습니까?`)) return;
                const res = await fetch(`/api/admin/students/${encodeURIComponent(username)}/reset-pw`, { method: 'POST' });
                const result = await res.json();
                alert(result.message);
            }

            async function deleteStudent(username) {
                if (!confirm(`${username} 학생을 삭제하시겠습니까?`)) return;
                await fetch(`/api/admin/students/${encodeURIComponent(username)}`, { method: 'DELETE' });
                fetchStudents();
            }

            window.onload = function() {
                updateSortIcons();
                fetchStudents();
            };
        </script>
    </body>
    </html>
    """

# ==============================================================================
# 3. 당일 신청 좌석표 인쇄 뷰어
# ==============================================================================
@app.get("/admin/print-seating-chart", response_class=HTMLResponse)
def print_seating_chart(class_name: str = ""):
    conn = get_db_connection()
    check_and_apply_resets(conn)
    cursor = conn.cursor()
    
    cursor.execute("SELECT room_name, rows_count, cols_count FROM class_configs WHERE class_name = ?", (class_name,))
    config = cursor.fetchone()
    room_name = config[0] if config else "301호"
    
    cursor.execute("SELECT title, rows_count, cols_count, grid_json FROM rooms WHERE room_name = ?", (room_name,))
    room_data = cursor.fetchone()
    
    if room_data:
        title, orig_rows, orig_cols, grid = room_data[0], room_data[1], room_data[2], json.loads(room_data[3])
    else:
        title = room_name
        grid = [[{"type": "seat", "id": f"{chr(65+r)}{c+1}"} for c in range(4)] for r in range(5)]

    actual_seat_count = sum(1 for row in grid for cell in row if cell.get("type") == "seat")

    cursor.execute("SELECT seat_id, name, username FROM seat_reservations WHERE class_name = ?", (class_name,))
    res_rows = cursor.fetchall()
    seat_to_name = {r[0]: r[1] for r in res_rows}
    reserved_user_set = set([r[2] for r in res_rows])

    cursor.execute("SELECT username, name, class_name FROM students")
    all_students = cursor.fetchall()
    conn.close()

    total_class_students = []
    unreserved_students = []
    for u_id, u_name, c_str in all_students:
        c_list = [c.strip() for c in c_str.split(',') if c.strip()]
        if class_name in c_list:
            total_class_students.append((u_id, u_name))
            if u_id not in reserved_user_set:
                unreserved_students.append((u_id, u_name))

    name_counts = Counter([u[1] for u in total_class_students])
    total_count = len(total_class_students)
    reserved_count = len(reserved_user_set)
    unreserved_count = len(unreserved_students)

    now = datetime.now()
    formatted_date = f"{now.year}년 {now.month}월 {now.day}일"

    table_rows_html = ""
    for row in grid:
        table_rows_html += "<tr>"
        for cell in row:
            if cell.get("type") == "seat":
                s_id = cell["id"]
                s_name = seat_to_name.get(s_id, "")
                name_display = f'<div style="font-size: 15px; font-weight: 800; color: #1e3a8a; margin-top: 4px;">{s_name}</div>' if s_name else '<div style="font-size: 11px; color: #94a3b8; margin-top: 6px;">| &nbsp;&nbsp;&nbsp; |</div>'
                bg_color = "background-color: #eff6ff;" if s_name else "background-color: #ffffff;"
                
                table_rows_html += f'''
                    <td style="border: 1px solid #475569; width: 85px; height: 58px; text-align: center; vertical-align: top; padding: 4px; {bg_color}">
                        <div style="font-size: 12px; font-weight: 700; color: #334155;">{s_id}</div>
                        {name_display}
                    </td>
                '''
            elif cell.get("type") == "aisle":
                table_rows_html += f'''
                    <td style="border-left: 1px dashed #94a3b8; border-right: 1px dashed #94a3b8; background-color: #f8fafc; font-size: 13px; font-weight: 800; color: #64748b; text-align: center; width: 35px;">
                        {cell["val"]}
                    </td>
                '''
            else:
                table_rows_html += '''
                    <td style="border: 1px solid #cbd5e1; background: linear-gradient(to top right, transparent calc(50% - 1px), #cbd5e1, transparent calc(50% + 1px)), linear-gradient(to bottom right, transparent calc(50% - 1px), #cbd5e1, transparent calc(50% + 1px)); width: 85px; height: 58px;"></td>
                '''
        table_rows_html += "</tr>"

    if unreserved_students:
        badges = []
        for u_id, u_name in unreserved_students:
            display_text = f"{u_name} ({u_id})" if name_counts[u_name] > 1 else u_name
            badges.append(f'<span style="display: inline-block; background: #fff; border: 1px solid #fecdd3; color: #e11d48; font-size: 11px; font-weight: bold; padding: 3px 8px; border-radius: 6px; margin: 2px 4px;">{display_text}</span>')
        unres_badges = "".join(badges)
    else:
        unres_badges = '<span style="color: #10b981; font-size: 12px; font-weight: bold;">모든 학생이 좌석 신청을 완료했습니다.</span>'

    return f"""
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <title>[김영편입 좌석표] {class_name} ({formatted_date})</title>
        <style>
            @import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard/dist/web/static/pretendard.css');
            * {{ box-sizing: border-box; font-family: 'Pretendard', sans-serif; }}
            body {{ margin: 0; padding: 20px; background-color: #f1f5f9; display: flex; flex-direction: column; align-items: center; }}
            .paper {{ background: white; width: 210mm; min-height: 297mm; padding: 15mm; box-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1); border-radius: 8px; display: flex; flex-direction: column; justify-content: space-between; }}
            @media print {{
                body {{ background: white; padding: 0; }}
                .paper {{ width: 100%; box-shadow: none; padding: 0; min-height: 100vh; }}
                .no-print {{ display: none !important; }}
            }}
        </style>
    </head>
    <body>
        <div class="no-print" style="margin-bottom: 20px; display: flex; gap: 10px;">
            <button onclick="window.print()" style="background-color: #2563eb; color: white; border: none; padding: 10px 24px; border-radius: 8px; font-weight: 700; cursor: pointer; font-size: 14px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">A4 인쇄 / PDF 다운로드</button>
            <button onclick="window.close()" style="background-color: #64748b; color: white; border: none; padding: 10px 18px; border-radius: 8px; font-weight: 600; cursor: pointer; font-size: 14px;">닫기</button>
        </div>

        <div class="paper">
            <div>
                <div style="background-color: #f59e0b; color: white; padding: 16px 20px; border-radius: 6px 6px 0 0; display: flex; justify-content: space-between; align-items: center;">
                    <div style="font-size: 22px; font-weight: 900;">[{class_name}] 좌석 배치표</div>
                    <div style="font-size: 15px; font-weight: 700; background: rgba(0,0,0,0.15); padding: 4px 10px; border-radius: 4px;">{room_name} ({actual_seat_count}석) - {formatted_date}</div>
                </div>

                <div style="display: flex; background-color: #cbd5e1; border: 1px solid #94a3b8; border-top: none; font-weight: 800; font-size: 13px;">
                    <div style="flex: 1; text-align: center; padding: 8px; letter-spacing: 12px; color: #1e293b;">칠 &nbsp;&nbsp; 판</div>
                    <div style="background-color: #475569; color: white; padding: 8px 16px;">출입문</div>
                </div>

                <div style="margin-top: 15px; display: flex; justify-content: center;">
                    <table style="border-collapse: collapse; margin: 0 auto;">
                        {table_rows_html}
                    </table>
                </div>
            </div>

            <div style="margin-top: 20px;">
                <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 10px 16px; display: flex; justify-content: space-between; align-items: center; font-size: 13px; font-weight: bold;">
                    <span style="color: #475569;">김영편입 좌석표 신청</span>
                    <div style="display: flex; gap: 16px;">
                        <span style="color: #334155;">총원: <b>{total_count}명</b></span>
                        <span style="color: #2563eb;">신청: <b>{reserved_count}명</b></span>
                        <span style="color: #e11d48;">미신청: <b>{unreserved_count}명</b></span>
                    </div>
                </div>

                <div style="margin-top: 8px; background-color: #fff1f2; border: 1px solid #ffe4e6; border-radius: 8px; padding: 10px 14px;">
                    <div style="font-size: 12px; font-weight: 800; color: #be123c; margin-bottom: 6px;">미신청 학생 명단 ({unreserved_count}명)</div>
                    <div style="line-height: 1.8;">
                        {unres_badges}
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

# ==============================================================================
# 4. 백엔드 API
# ==============================================================================
@app.post("/api/login")
def api_login(username: str = Form(...), password: str = Form(...)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT name, class_name FROM students WHERE username = ? AND password = ?", (username, password))
    user = cursor.fetchone()
    conn.close()
    if user:
        raw_classes = [c.strip() for c in user[1].split(",") if c.strip() and c.strip() != "-"]
        return {"success": True, "username": username, "name": user[0], "class_list": raw_classes if raw_classes else ["기본반"]}
    return {"success": False, "message": "아이디 또는 비밀번호가 일치하지 않습니다."}

@app.post("/api/change-password")
def api_change_password(username: str = Form(...), current_password: str = Form(...), new_password: str = Form(...)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM students WHERE username = ? AND password = ?", (username, current_password))
    user = cursor.fetchone()
    if not user:
        conn.close()
        return {"success": False, "message": "현재 비밀번호가 일치하지 않습니다."}
        
    cursor.execute("UPDATE students SET password = ? WHERE username = ?", (new_password, username))
    conn.close()
    return {"success": True, "message": "비밀번호가 성공적으로 변경되었습니다."}

@app.post("/api/admin/students/{username}/reset-pw")
def api_reset_password(username: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE students SET password = '1234' WHERE username = ?", (username,))
    conn.close()
    return {"success": True, "message": f"{username} 학생의 비밀번호가 1234로 초기화되었습니다."}

@app.get("/api/seats")
def api_seats(username: str, class_name: str):
    conn = get_db_connection()
    check_and_apply_resets(conn)
    cursor = conn.cursor()
    
    cursor.execute("SELECT room_name, rows_count, cols_count FROM class_configs WHERE class_name = ?", (class_name,))
    config = cursor.fetchone()
    room_name = config[0] if config else "301호"
    
    cursor.execute("SELECT title, rows_count, cols_count, grid_json FROM rooms WHERE room_name = ?", (room_name,))
    room_data = cursor.fetchone()
    
    if room_data:
        title, orig_rows, orig_cols, grid = room_data[0], room_data[1], room_data[2], json.loads(room_data[3])
    else:
        title = room_name
        grid = [[{"type": "seat", "id": f"{chr(65+r)}{c+1}"} for c in range(4)] for r in range(5)]
        
    cols = len(grid[0]) if grid else 4

    cursor.execute("SELECT seat_id, username FROM seat_reservations WHERE class_name = ?", (class_name,))
    reservations = dict(cursor.fetchall())
    conn.close()

    total_seats = 0
    for r in grid:
        for cell in r:
            if cell.get("type") == "seat":
                total_seats += 1
                assigned_user = reservations.get(cell["id"])
                if assigned_user == username:
                    cell["status_class"] = "mine"
                elif assigned_user:
                    cell["status_class"] = "occupied"
                else:
                    cell["status_class"] = "empty"
                    
    return {"room_name": room_name, "room_title": title, "cols": cols, "total_seats": total_seats, "grid": grid}

@app.post("/api/reserve")
def api_reserve(username: str = Form(...), name: str = Form(...), class_name: str = Form(...), seat_id: str = Form(...)):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("BEGIN IMMEDIATE;")
        check_and_apply_resets(conn)
        
        cursor.execute("SELECT username FROM seat_reservations WHERE class_name = ? AND seat_id = ?", (class_name, seat_id))
        target = cursor.fetchone()
        
        if target and target[0] and target[0] != username:
            conn.execute("ROLLBACK;")
            conn.close()
            return {"success": False, "message": "방금 다른 학생이 먼저 신청한 좌석입니다!"}
            
        cursor.execute("DELETE FROM seat_reservations WHERE class_name = ? AND username = ?", (class_name, username))
        cursor.execute("INSERT OR REPLACE INTO seat_reservations (class_name, seat_id, username, name) VALUES (?, ?, ?, ?)", (class_name, seat_id, username, name))
        conn.execute("COMMIT;")
        conn.close()
        return {"success": True}
    except Exception:
        try: conn.execute("ROLLBACK;")
        except: pass
        conn.close()
        return {"success": False, "message": "좌석 신청 처리 중 일시적인 경합이 발생했습니다. 다시 시도해주세요."}

@app.get("/api/admin/students")
def get_students(keyword: str = "", sort_by: str = "username", order: str = "asc", class_filter: str = ""):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    valid_cols = {"username": "username", "name": "name", "class_name": "class_name"}
    col_sql = valid_cols.get(sort_by, "username")
    order_sql = "DESC" if order.lower() == "desc" else "ASC"

    query = f"SELECT username, name, class_name FROM students WHERE 1=1"
    params = []
    if keyword:
        query += " AND (username LIKE ? OR name LIKE ? OR class_name LIKE ?)"
        params.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
        
    query += f" ORDER BY {col_sql} {order_sql}"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    cursor.execute("SELECT COUNT(*) FROM students")
    total = cursor.fetchone()[0]
    conn.close()

    filter_classes = [c.strip() for c in class_filter.split(",") if c.strip()]
    student_list = []
    for r in rows:
        u_id, u_name, c_str = r[0], r[1], r[2]
        student_c_list = [c.strip() for c in c_str.split(",") if c.strip()]
        if filter_classes and not any(fc in student_c_list for fc in filter_classes):
            continue
        student_list.append({"username": u_id, "name": u_name, "class_name": c_str})

    return {"total": total, "filtered_total": len(student_list), "students": student_list}

@app.post("/api/admin/students/clear-all")
def clear_all_students():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM students")
    cursor.execute("DELETE FROM seat_reservations")
    cursor.execute("DELETE FROM class_configs")
    conn.close()
    return {"success": True, "message": "모든 학생 명단 및 좌석 신청 데이터가 성공적으로 초기화되었습니다."}

@app.post("/api/admin/students/add-manual")
def add_student_manual(username: str = Form(...), name: str = Form(...), class_names: str = Form(...)):
    username, name, class_names = username.strip(), name.strip(), class_names.strip()
    if not username or not name:
        return {"success": False, "message": "아이디와 이름을 모두 입력해 주세요."}

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM students WHERE username = ?", (username,))
    if cursor.fetchone():
        conn.close()
        return {"success": False, "message": f"이미 존재하는 아이디({username})입니다."}

    cursor.execute("INSERT INTO students (username, name, class_name, password, status) VALUES (?, ?, ?, '1234', 'active')", (username, name, class_names))

    for c in class_names.split(','):
        c_clean = c.strip()
        if c_clean and c_clean != '-':
            cursor.execute("INSERT OR IGNORE INTO class_configs (class_name, room_name, rows_count, cols_count) VALUES (?, '301호', 0, 0)", (c_clean,))

    conn.close()
    return {"success": True, "message": f"학생 [{name}({username})]이 성공적으로 추가되었습니다."}

@app.delete("/api/admin/students/{username}")
def delete_student(username: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM students WHERE username = ?", (username,))
    cursor.execute("DELETE FROM seat_reservations WHERE username = ?", (username,))
    conn.close()
    return {"success": True}

@app.get("/api/admin/classes")
def get_classes():
    conn = get_db_connection()
    check_and_apply_resets(conn)
    cursor = conn.cursor()
    
    cursor.execute("SELECT room_name, title, rows_count, cols_count, grid_json FROM rooms ORDER BY room_name")
    rooms = cursor.fetchall()
    
    rooms_list = []
    rooms_dict = {}
    for r in rooms:
        r_name, r_title, r_rows, r_cols, r_grid_str = r[0], r[1], r[2], r[3], r[4]
        try:
            grid_data = json.loads(r_grid_str) if r_grid_str else []
            seat_cnt = sum(1 for row in grid_data for cell in row if cell.get("type") == "seat")
        except Exception:
            seat_cnt = 0
            
        rooms_dict[r_name] = {"rows": r_rows, "cols": r_cols, "total_seats": seat_cnt}
        rooms_list.append({
            "room_name": r_name, 
            "title": r_title, 
            "rows_count": r_rows, 
            "cols_count": r_cols,
            "total_seats": seat_cnt
        })
    
    cursor.execute("SELECT class_name, room_name, rows_count, cols_count, reset_datetime FROM class_configs ORDER BY class_name")
    classes = cursor.fetchall()
    
    cursor.execute("SELECT username, name, class_name FROM students")
    all_students = cursor.fetchall()
    
    cursor.execute("SELECT class_name, username FROM seat_reservations")
    all_reservations = cursor.fetchall()
    conn.close()

    class_students_map = {}
    for u_id, u_name, c_str in all_students:
        for c in c_str.split(','):
            c_clean = c.strip()
            if c_clean and c_clean != '-':
                class_students_map.setdefault(c_clean, []).append((u_id, u_name))
                
    class_reserved_map = {}
    for c_name, u_id in all_reservations:
        class_reserved_map.setdefault(c_name, set()).add(u_id)

    class_list = []
    for c in classes:
        c_name, r_name, r_cnt, c_cnt, reset_dt = c[0], c[1], c[2], c[3], c[4]
        if (r_cnt == 0 or c_cnt == 0) and r_name in rooms_dict:
            r_cnt = rooms_dict[r_name]["rows"]
            c_cnt = rooms_dict[r_name]["cols"]
            
        st_list = class_students_map.get(c_name, [])
        total_st = len(st_list)
        reserved_set = class_reserved_map.get(c_name, set())
        reserved_st = len([s for s in st_list if s[0] in reserved_set])
        
        class_list.append({
            "class_name": c_name,
            "room_name": r_name,
            "rows_count": r_cnt or 0,
            "cols_count": c_cnt or 0,
            "reset_datetime": reset_dt or "",
            "total_students": total_st,
            "reserved_count": reserved_st,
            "unreserved_count": max(0, total_st - reserved_st)
        })
        
    return {
        "classes": class_list,
        "rooms": rooms_list
    }

@app.get("/api/admin/reserved-students")
def get_reserved_students(class_name: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT seat_id, username, name FROM seat_reservations WHERE class_name = ? ORDER BY seat_id", (class_name,))
    reserved_list = [{"seat_id": r[0], "username": r[1], "name": r[2]} for r in cursor.fetchall()]
    conn.close()
    return {"class_name": class_name, "reserved": reserved_list}

@app.get("/api/admin/unreserved-students")
def get_unreserved_students(class_name: str):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username, name, class_name FROM students")
    all_students = cursor.fetchall()
    
    cursor.execute("SELECT username FROM seat_reservations WHERE class_name = ?", (class_name,))
    reserved_users = set([r[0] for r in cursor.fetchall()])
    conn.close()

    total_class_students = []
    unreserved_list = []
    for u_id, u_name, c_str in all_students:
        c_list = [c.strip() for c in c_str.split(',') if c.strip()]
        if class_name in c_list:
            total_class_students.append(u_name)
            if u_id not in reserved_users:
                unreserved_list.append({"username": u_id, "name": u_name})
                
    name_counts = Counter(total_class_students)
    for item in unreserved_list:
        item["is_duplicate_name"] = name_counts[item["name"]] > 1
            
    return {"class_name": class_name, "unreserved": unreserved_list}

@app.post("/api/admin/classes/update")
def update_class_config(
    class_name: str = Form(...), 
    room_name: str = Form(...), 
    rows_count: int = Form(0), 
    cols_count: int = Form(0),
    reset_datetime: str = Form("")
):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO class_configs (class_name, room_name, rows_count, cols_count, reset_datetime)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(class_name) DO UPDATE SET 
            room_name = excluded.room_name,
            rows_count = excluded.rows_count,
            cols_count = excluded.cols_count,
            reset_datetime = excluded.reset_datetime
    """, (class_name, room_name, rows_count, cols_count, reset_datetime))
    check_and_apply_resets(conn)
    conn.close()
    return {"success": True, "message": f"[{class_name}] 설정이 성공적으로 저장되었습니다."}

@app.post("/api/admin/classes/reset-now")
def reset_class_now(class_name: str = Form(...)):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM seat_reservations WHERE class_name = ?", (class_name,))
    conn.close()
    return {"success": True, "message": f"[{class_name}] 반의 좌석 신청 내역이 즉시 초기화되었습니다."}

@app.post("/api/admin/students/upload")
async def upload_students_file(file: UploadFile = File(...)):
    contents = await file.read()
    df = parse_student_file(contents)

    if df is None or '아이디' not in df.columns or '이름' not in df.columns:
        return {"message": "파일에서 '아이디'와 '이름' 컬럼을 찾을 수 없습니다."}

    class_col = '강좌명' if '강좌명' in df.columns else ('반명' if '반명' in df.columns else None)

    if class_col:
        grouped = df.groupby(['아이디', '이름'])[class_col].apply(
            lambda s: ', '.join(dict.fromkeys(s.dropna().astype(str)))
        ).reset_index()
    else:
        grouped = df[['아이디', '이름']].drop_duplicates()
        grouped['class_name'] = '-'

    conn = get_db_connection()
    cursor = conn.cursor()
    
    unique_classes = set()
    for _, row in grouped.iterrows():
        u_id = str(row['아이디']).strip()
        u_name = str(row['이름']).strip()
        c_name = str(row[class_col if class_col else 'class_name']).strip()
        
        for c in c_name.split(','):
            if c.strip() and c.strip() != '-':
                unique_classes.add(c.strip())

        cursor.execute("""
            INSERT INTO students (username, name, class_name, password, status)
            VALUES (?, ?, ?, '1234', 'active')
            ON CONFLICT(username) DO UPDATE SET
                name = excluded.name,
                class_name = excluded.class_name
        """, (u_id, u_name, c_name))
        
    for c in unique_classes:
        cursor.execute("INSERT OR IGNORE INTO class_configs (class_name, room_name, rows_count, cols_count) VALUES (?, '301호', 0, 0)", (c,))

    conn.close()
    return {"message": f"총 {len(grouped)}명의 학생과 {len(unique_classes)}개 반 목록이 등록되었습니다!"}

@app.post("/api/admin/rooms/upload")
async def upload_room_excel(file: UploadFile = File(...)):
    contents = await file.read()
    parsed_rooms = parse_seat_excel(contents)
    
    if not parsed_rooms:
        return {"message": "좌석표 엑셀 파일에서 A1, B1 형태의 좌석을 찾을 수 없습니다."}
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    first_room = parsed_rooms[0]
    
    for r in parsed_rooms:
        cursor.execute("""
            INSERT INTO rooms (room_name, title, rows_count, cols_count, grid_json)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(room_name) DO UPDATE SET
                title = excluded.title,
                rows_count = excluded.rows_count,
                cols_count = excluded.cols_count,
                grid_json = excluded.grid_json
        """, (r["room_name"], r["title"], r["rows_count"], r["cols_count"], json.dumps(r["grid"])))
        
        cursor.execute("""
            UPDATE class_configs 
            SET rows_count = ?, cols_count = ?
            WHERE room_name = ?
        """, (r["rows_count"], r["cols_count"], r["room_name"]))

    cursor.execute("""
        UPDATE class_configs 
        SET room_name = ?, rows_count = ?, cols_count = ?
        WHERE rows_count = 0 OR cols_count = 0
    """, (first_room["room_name"], first_room["rows_count"], first_room["cols_count"]))
        
    conn.close()
    room_summary = ", ".join([f"{r['room_name']} ({r['total_seat_count']}석)" for r in parsed_rooms])
    return {"message": f"총 {len(parsed_rooms)}개 강의실 좌석표가 정확히 등록되었습니다!\n({room_summary})"}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("app:app", host="0.0.0.0", port=port)
